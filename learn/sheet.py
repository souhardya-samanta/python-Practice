import openpyxl as xl
wb = xl.load_workbook("product.xlsx")
sheet = wb['Sheet1']
for units in range(2, 45):
    total = sheet.cell(units, 7)
    rounded_data = round(total.value)
    rounded_data_cell = sheet.cell(units, 8)
    rounded_data_cell.value = rounded_data
wb.save("new.xlsx")

